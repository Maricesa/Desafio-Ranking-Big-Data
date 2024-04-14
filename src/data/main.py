import pandas as pd
import openpyxl
import glob
import os

#Caminho dos arquivos excel
folder_path = 'DESAFIO_RANKING\\src\\data\\raw\\'

#Carrega os arquivos
book = pd.read_excel(folder_path)

#lista todos os arquivos de excel
book = glob.glob(os.path.join(folder_path,'*.xlsx'))
for excel_file in book:
 print(excel_file.filename)

if not book:
    print('Nenhum arquivo compatível encontrado!')

else:
    #tabela na memória para guardar os arquivos
    dfs = [folder_path]
    
    for excel_file in book:
        try:
            #Salva o arquivo a ser lido numa dataframe temporária
            df_temp = openpyxl.load_workbook(excel_file)

            #lê as abas do arquivo
            df_temp = pd.read_excel(excel_file,sheet_name=[])          
                
            #cria uma nova planilha para escrita dos dados
            file_ready = openpyxl.workbook()  

            #carrega o arquivo de origem para copiar os dados
            file_o1 = openpyxl.load_workbook(df_temp)
            file_i = file_o1.active

            #Carrega o arquivo de destino
            file_d2 = openpyxl.load_workbook(file_ready)
            file_i2 = file_d2.active

            # Especificação do intervalo de células de origem (por exemplo, A1:C10)
            intervalo_origem = file_o1['E6:W1']

            # Especificação do intervalo de células de destino (por exemplo, D1:F10)
            intervalo_destino = file_d2['A1:A32']

            #Copia os valores do intervalo de origem para o intervalo de destino
            for row_origem, row_destino in zip(intervalo_origem, intervalo_destino):
                for cell_origem, cell_destino in zip(row_origem,row_destino):
                    cell_destino.value = cell_origem.value
            
            #Salva o arquivo de destino
            file_ready.save('Planilha final.xlsx')
            print("Os dados foram copiados com sucesso")

            print (file_ready)

            dfs.append(file_ready)
           
        except Exception as e:
             print(f'Erro ao ler o arquivo {excel_file}: {e} ')


